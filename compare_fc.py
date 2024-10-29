import pandas as pd
import numpy as np
from openpyxl import load_workbook


def read_data(*paths):
    dfs = []
    column_map = {
        1: 'MZmine 3',
        2: 'PeakDetective',
        3: 'Peakonly',
        4: 'QuanFormer'
    }

    for i, path in enumerate(paths):
        if path is not None:
            if i == 0:
                df = pd.read_csv(path, usecols=['Compound Name', 'Compound concentration ratio', 'Fold change'])
                df = df.rename(columns={'Fold change': 'Ground Truth'})
            else:
                df = pd.read_csv(path, usecols=['Compound Name', 'fc'])
                df = df.rename(columns={'fc': column_map[i]})
            dfs.append(df)
    return tuple(dfs)


def read_data_CV(*paths):
    dfs = []
    column_map = {
        1: 'MZmine 3',
        2: 'PeakDetective',
        3: 'Peakonly',
        4: 'QuanFormer'
    }

    for i, path in enumerate(paths):
        if path is not None:
            if i == 0:
                df = pd.read_csv(path, usecols=['Compound Name'])
            else:
                df = pd.read_csv(path, usecols=['Compound Name', 'CV-A', 'CV-B'])
                df = df.rename(columns={'CV-A': column_map[i] + ' (SA)', 'CV-B': column_map[i] + ' (SB)'})
            dfs.append(df)
    return tuple(dfs)


def merge_data(*dfs):
    merge_df = dfs[0]
    for df in dfs[1:]:
        merge_df = pd.merge(merge_df, df, on='Compound Name', how='outer')
    return merge_df


def swap_columns_and_rename(merged_df, col1, col2):
    temp_col = merged_df[col1].copy()

    merged_df[col1] = merged_df[col2]
    merged_df[col2] = temp_col

    col_names = merged_df.columns.tolist()
    col_names[col_names.index(col1)], col_names[col_names.index(col2)] = col_names[col_names.index(col2)], col_names[
        col_names.index(col1)]
    merged_df.columns = col_names


def plot_boxplot(*paths, output):
    dfs = read_data(*paths)
    merged_df = merge_data(*dfs)

    merged_df = merged_df.drop(columns=['Compound Name'])

    swap_columns_and_rename(merged_df, 'Compound concentration ratio', 'Ground Truth')

    count = merged_df.shape[1] - 1
    merged_df.iloc[:, -count:] = merged_df.iloc[:, -count:].applymap(
        lambda x: np.log2(x) if pd.api.types.is_numeric_dtype(type(x)) and x > 0 else np.nan
    )

    custom_order = ['1/16', '1/4', '1/2', '1/1', '2/1', '4/1', '16/1']
    merged_df['Compound concentration ratio'] = pd.Categorical(merged_df['Compound concentration ratio'],
                                                                categories=custom_order, ordered=True)
    df_sorted = merged_df.sort_values(by='Compound concentration ratio')

    column_map = {
        0: 'Ground Truth',
        1: 'MZmine 3',
        2: 'PeakDetective',
        3: 'Peakonly',
        4: 'QuanFormer'
    }

    for i in range(count):
        col = df_sorted.iloc[:, [0, i + 1]]
        if 'Shift' in output and i == 3:
            i = 4
        name = f'{output}_{column_map[i]}.xlsx'
        write_to_excel(col, name)


def write_to_excel(df, output):
    with pd.ExcelWriter(output) as writer:
        grouped = df.groupby('Compound concentration ratio')

        for name, group in grouped:
            group.to_excel(writer, sheet_name=name.replace('/', '_'), index=False)

    wb = load_workbook(output)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws['A1'] = sheet.replace('_', '/')

    wb.save(output)


def plot_boxplotCV(*paths, output):
    dfs = read_data_CV(*paths)
    merged_df = merge_data(*dfs)
    merged_df.to_excel(output, index=False)


if __name__ == '__main__':
    # Figure 3
    path0 = 'data/FeatureQE.csv'
    path1 = 'data/MZMINE3/QE.csv'
    path2 = 'data/PEAKDETECTIVE/peakdetective_QE.csv'
    path3 = 'data/peakonly/QE.csv'
    path4 = 'data/Mine/QE/post-my-qe.csv'

    plot_boxplot(path0, path1, path2, path3, path4, output='QE')
    #
    path0 = 'data/FeatureQEShift.csv'
    path1 = 'data/MZMINE3/QEShift.csv'
    path2 = 'data/PEAKDETECTIVE/peakdetective_QEShift.csv'
    path3 = None
    path4 = 'data/Mine/QEShift/post-my-qe.csv'
    plot_boxplot(path0, path1, path2, path3, path4, output='QEShift')
    #
    path0 = 'data/FeatureTOF.csv'
    path1 = 'data/MZMINE3/TOF.csv'
    path2 = 'data/PEAKDETECTIVE/peakdetective_TOF.csv'
    path3 = 'data/peakonly/TOF.csv'
    path4 = 'data/Mine/TOF/post-my-tof.csv'

    plot_boxplot(path0, path1, path2, path3, path4, output='TOF')
    #
    path0 = 'data/FeatureTOFShift.csv'
    path1 = 'data/MZMINE3/TOFShift.csv'
    path2 = 'data/PEAKDETECTIVE/peakdetective_TOFShift.csv'
    path3 = None
    path4 = 'data/Mine/TOFShift/post-my-tof.csv'
    plot_boxplot(path0, path1, path2, path3, path4, output='TOFShift')

    # # Figure4
    # path0 = 'data/FeatureQE.csv'
    # path1 = 'data/MZMINE3/QE.csv'
    # path2 = 'data/PEAKDETECTIVE/peakdetective_QE.csv'
    # path3 = 'data/peakonly/QE.csv'
    # path4 = 'data/Mine/QE/post-my-qe.csv'
    # plot_boxplotCV(path0, path1, path2, path3, path4, output='QE_CV1.xlsx')
    #
    # path0 = 'data/FeatureQEShift.csv'
    # path1 = 'data/MZMINE3/QEShift.csv'
    # path2 = 'data/PEAKDETECTIVE/peakdetective_QEShift.csv'
    # path3 = None
    # path4 = 'data/Mine/QEShift/post-my-qe.csv'
    # plot_boxplotCV(path0, path1, path2, path3, path4, output='QEShift_CV1.xlsx')
    #
    # path0 = 'data/FeatureTOF.csv'
    # path1 = 'data/MZMINE3/TOF.csv'
    # path2 = 'data/PEAKDETECTIVE/peakdetective_TOF.csv'
    # path3 = 'data/peakonly/TOF.csv'
    # path4 = 'data/Mine/TOF/post-my-tof.csv'
    # plot_boxplotCV(path0, path1, path2, path3, path4, output='TOF_CV1.xlsx')
    #
    # path0 = 'data/FeatureTOFShift.csv'
    # path1 = 'data/MZMINE3/TOFShift.csv'
    # path2 = 'data/PEAKDETECTIVE/peakdetective_TOFShift.csv'
    # path3 = None
    # path4 = 'data/Mine/TOFShift/post-my-tof.csv'
    # plot_boxplotCV(path0, path1, path2, path3, path4, output='TOFShift_CV1.xlsx')
