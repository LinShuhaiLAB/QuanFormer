import pandas as pd
import numpy as np


def plot_boxplot(path0, path1, path2, path3, Name):
    df0 = pd.read_csv(path0, usecols=['Compound Name', 'Compound concentration ratio', 'Fold change'])
    df1 = pd.read_csv(path1, usecols=['Compound Name', 'fc'])
    df2 = pd.read_csv(path2, usecols=['Compound Name', 'fc'])
    df3 = pd.read_csv(path3, usecols=['Compound Name', 'fc'])

    merged_df = pd.merge(df0, df1, on='Compound Name', how='outer')
    merged_df = pd.merge(merged_df, df2, on='Compound Name', how='outer')
    merged_df = pd.merge(merged_df, df3, on='Compound Name', how='outer')
    merged_df.columns = ['Compound Name', 'Compound concentration ratio', 'Ground Truth', 'MZmine 3', 'PeakDetective', 'PeakFormer']

    merged_df = merged_df.drop(columns=['Compound Name'])

    temp_col = merged_df[merged_df.columns[0]]
    merged_df[merged_df.columns[0]] = merged_df[merged_df.columns[1]]
    merged_df[merged_df.columns[1]] = temp_col

    merged_df.iloc[:, -4:] = merged_df.iloc[:, -4:].applymap(
        lambda x: np.log2(x) if pd.api.types.is_numeric_dtype(type(x)) and x > 0 else np.nan
    )

    df_sorted = merged_df.sort_values(by='Compound concentration ratio')

    with pd.ExcelWriter(Name) as writer:
        grouped = df_sorted.groupby('Compound concentration ratio')

        for name, group in grouped:
            group.to_excel(writer, sheet_name=name.replace('/', '_') , index=False)

    from openpyxl import load_workbook
    wb = load_workbook(Name)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws['A1'] = sheet.replace('_', '/')

    wb.save(Name)


def plot_boxplotCV(path0, path1, path2, path3, Name):
    df0 = pd.read_csv(path0, usecols=['Compound Name'])
    df1 = pd.read_csv(path1, usecols=['Compound Name', 'CV-A', 'CV-B'])
    df2 = pd.read_csv(path2, usecols=['Compound Name', 'CV-A', 'CV-B'])
    df3 = pd.read_csv(path3, usecols=['Compound Name', 'CV-A', 'CV-B'])

    merged_df = pd.merge(df0, df1, on='Compound Name', how='outer')
    merged_df = pd.merge(merged_df, df2, on='Compound Name', how='outer')
    merged_df = pd.merge(merged_df, df3, on='Compound Name', how='outer')
    merged_df.columns = ['Compound Name', 'MZmine 3 (SA)', 'MZmine 3 (SB)', 'PeakDetective (SA)', 'PeakDetective (SB)',
                         'PeakFormer (SA)', 'PeakFormer (SB)']
    merged_df.to_excel(Name, index=False)


# path0 = 'data/FeatureTOF.csv'
# path1 = 'data/MZMINE3/TOF.csv'
# path2 = 'data/PEAKDETECTIVE/peakdetective_TOF.csv'
# path3 = 'data/Mine/TOF/post-my-tof.csv'
# Name = 'TOF_FC.xlsx'
# plot_boxplot(path0, path1, path2, path3, Name)

path0 = 'data/FeatureQE.csv'
path1 = 'data/MZMINE3/QE.csv'
path2 = 'data/PEAKDETECTIVE/peakdetective_QE.csv'
path3 = 'data/Mine/QE/post-my-qe.csv'
Name = 'QE_CV.xlsx'
plot_boxplotCV(path0, path1, path2, path3, Name)

